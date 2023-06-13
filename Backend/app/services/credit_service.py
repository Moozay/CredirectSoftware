from datetime import date, datetime
from app.models.prospect_model import Prospect
from app.services.prospect_service import ProspectService
from app.services.user_service import UserService
from app.services.coemp_service import CoempService
from app.schemas.demande_credit_schema import DemandeCreditCreate
from app.schemas.prospect_schema import ProspectRecord
from datetime import date
from app.schemas.credit_schema import CreditCreate, CreditBulkUpdate, CreditUpdateHonnaires, CreditUpdateStatus, CreditOut, CreditDisplay
from uuid import UUID, uuid4
from typing import List, Optional
from beanie.operators import In
from fastapi import Request
from fastapi.responses import FileResponse
from fastapi import HTTPException
from app.models.credit_model import Credit
from fastapi.templating import Jinja2Templates
from app.models.demande_credit_model import DemandeCredit
import pymongo
from docxtpl import DocxTemplate
import os
import locale
from openpyxl.styles.borders import Border, Side
import openpyxl
from copy import copy
import pdfkit


class CreditService:
    @staticmethod
    async def create_credit(credit: CreditCreate):
        credit_in = Credit(
            credit_id=credit.credit_id,
            type_credit=credit.type_credit,
            montant=credit.montant,
            duree_credit=credit.duree_credit,
            frequence=credit.frequence,
            mensualite=credit.mensualite,
            taux=credit.taux,
            taux_demande=credit.taux_demande,
            franchise=credit.franchise,
            taux_endt=credit.taux_endt,
            teg=credit.teg,
            qot_financement=credit.qot_financement,
            objet_credit=credit.objet_credit,
            nature_bien=credit.nature_bien,
            etat_bien=credit.etat_bien,
            usage_bien=credit.usage_bien,
            montant_acte=credit.montant_acte,
            montant_travaux=credit.montant_travaux,
            montant_venal=credit.montant_venal,
            adresse_bien=credit.adresse_bien,
            superficie=credit.superficie,
            prospect_id=credit.prospect_id,
            titre_foncier=credit.titre_foncier,
            garanties=credit.garanties,
            statusCredit=credit.statusCredit,
            commentaires=credit.commentaires,
            promoteur=credit.promoteur,
            promoteur_nom=credit.promoteur_nom,
            banque_envoye=credit.banque_envoye,
            date=datetime.utcnow().isoformat(),
            status_honnaires={
                "hc": False,
                "hb": False,
                "date_hb": None,
                "date_hc":  None
            },
            coemp=credit.coemp,
            prospect_revenue=credit.prospect_revenue,
            engagements_bancaires=credit.engagements_bancaires,
            agent_id=credit.agent_id
        )
        await credit_in.save()
        return credit_in

    @staticmethod
    async def add_credit(data: CreditCreate):
        prospect = await Prospect.find_one(Prospect.prospect_id == data.prospect_id)
        credit = await CreditService.create_credit(data)
        credits = prospect.credits
        credits.append(credit.credit_id)
        await prospect.update({"$set": {"credits": credits}})

        return credit

    @staticmethod
    async def update_credit_status(id: UUID, data: CreditUpdateStatus) -> Credit:

        credit = await Credit.find_one(Credit.credit_id == id)
        if not Credit:
            raise pymongo.errors.OperationFailure("Credit not found")
        if data.statusCredit == "Débloqué" and credit.statusCredit != "Débloqué":
            data.date_debloque = datetime.utcnow()
        await credit.update({"$set": data.dict(exclude_unset=True)})
        return credit

    @staticmethod
    async def update_credit_honnaires(id: UUID, data: CreditUpdateHonnaires) -> Credit:

        credit = await Credit.find_one(Credit.credit_id == id)
        if not Credit:
            raise pymongo.errors.OperationFailure("Credit not found")
        if data.status_honnaires["hb"] == True:
            data.status_honnaires = {
                **data.status_honnaires, "date_hb": datetime.utcnow().isoformat()}

        if data.status_honnaires["hc"] == True:
            data.status_honnaires = {
                **data.status_honnaires, "date_hc": datetime.utcnow().isoformat()}

        await credit.update({"$set": data.dict(exclude_unset=True)})
        return credit

    @staticmethod
    async def update_credit_bulk(id: UUID, data: CreditBulkUpdate) -> Credit:
        credit = await Credit.find_one(Credit.credit_id == id)
        if not Credit:
            raise pymongo.errors.OperationFailure("Credit not found")

        await credit.update({"$set": data.dict(exclude_unset=True)})
        return credit

    @staticmethod
    async def get_dc(request: Request, credit_id: UUID):
        templates = Jinja2Templates(directory="app/static")
        context = {'request': request, 'date': date.today()}
        credit = await CreditService.get_credit_by_id(credit_id)
        prospect = await ProspectService.get_prospect_by_id(credit.prospect_id)
        agent = await UserService.get_user_by_id(prospect.agent_id)
        if prospect.coemp_id is not None:
            co_emprunteur = await CoempService.get_coemp_by_id(credit.coemp[0])
        else:
            co_emprunteur = 0
        context = {**context, 'credit': credit, 'prospect': prospect,
                   'agent': agent, 'co_emprunteur': co_emprunteur}
        html_file = "dc.html"
        if credit.type_credit == "consommation":
            html_file = "dc-consommation.html"
        else:
            credit.garanties = " , ".join(map(str, credit.garanties))
        return templates.TemplateResponse(html_file, context)

    @staticmethod
    async def create_dc(credit_id: UUID):
        filelocation = "app/temp/" + str(credit_id) + '.pdf'
        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')
        options = {
            'page-size': 'Letter',
            'margin-top': '0.5in',
            'margin-right': '0.1in',
            'margin-bottom': '0.6in',
            'margin-left': '0.1in',
            'encoding': "UTF-8",
            'footer-center': 'CREDIRECT.MA 73 Bd. Anfa, 5 ETG angle 1 rue Clos de Provence Casablanca - Maroc',
            'footer-font-size': '7',
            'footer-font-name': 'Georgia',
            'footer-right': '[page] of [topage]',
            'no-outline': None
        }
        html_url = "192.168.11.200:8000/api/v1/credits/demandecredit/" + \
            str(credit_id)
        pdfkit.from_url(html_url, output_path=filelocation,
                        options=options, configuration=config)

    @staticmethod
    async def download_dc(request: Request, credit_id: UUID):
        await CreditService.create_dc(credit_id)
        parentPath = os.path.dirname(__file__)
        path = "/../temp/"
        filePath = parentPath + path + str(credit_id) + '.pdf'
        st_abs = os.path.join(parentPath, filePath)
        filename = os.path.basename(st_abs)
        headers = {'content-Dispostion': f'attachment; filename="{filename}"'}
        return FileResponse(st_abs, headers=headers, media_type='application/pdf')

    @staticmethod
    async def switch_mandat(montant: float):
        frais = 0
        if montant >= 300000:
            frais = (3/100)*montant
            return frais
        elif montant < 300000:
            frais = 5000
            return frais
        elif montant < 100000:
            frais = 4000
            return frais
        elif montant < 75000:
            frais = 3500
            return frais
        elif montant < 50000:
            frais = 3000
            return frais
        elif montant < 40000:
            frais = 2000
            return frais
        elif montant < 30000:
            frais = 1500
            return frais
        elif montant < 20000:
            frais = 1000
            return frais

    @staticmethod
    async def change_string_float(montant: str):
        if montant == '':
            return 0
        montant = montant.replace(" ", "")
        montant = montant.replace(",", ".")
        montant = float(montant)
        return montant

    @staticmethod
    async def change_float_string(frais: float):
        frais = "{:,.2f}".format(frais)
        main, fractional = frais.split(".")
        main = main.replace(",", " ")
        frais = main + "," + fractional
        return frais

    @staticmethod
    async def download_facture(start, end, bank):
        if bank == "SOFAC":
            await CreditService.generate_sofac_facture(start, end, bank)
            fileLocation = "app/temp/facture_sofac.xlsx"
        elif bank == "CDM":
            await CreditService.generate_cdm_facture(start, end, bank)
            fileLocation = "app/temp/facture_cdm.xlsx"
        elif bank == "BMCI":
            await CreditService.generate_bmci_facture(start, end, bank)
            fileLocation = "app/temp/facture_bmci.xlsx"
        elif bank == "SGMB":
            await CreditService.generate_sgmb_facture(start, end, bank)
            fileLocation = "app/temp/facture_sgmb.xlsx"
        elif bank == "BP":
            await CreditService.generate_bp_facture(start, end, bank)
            fileLocation = "app/temp/facture_bp.xlsx"
        headers = {'Content-Disposition': 'attachment; filename="Book.xlsx"'}
        return FileResponse(fileLocation, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @staticmethod
    async def generate_sofac_facture(start, end, bank):
        wb = openpyxl.load_workbook("app/static/facture_sofac.xlsx")
        ws = wb["Sheet1"]
        start_row = 14
        demandes = await Credit.find_many({"banque": "SOFAC", "statusCredit": "Débloqué", "status_honnaires.hb": False, "date_debloque": {
            "$gte": datetime.strptime(start, '%d/%m/%Y'),
            "$lte": datetime.strptime(end, '%d/%m/%Y'),
        }}).to_list()
        if demandes.__len__() == 0 :
            raise HTTPException(status_code=404, detail="Aucun enregistrement trouvé")
        today_date = datetime.now().strftime("%m/%d/%Y")
        ws.cell(row=10, column=5).value = today_date
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        selected_month = datetime.strptime(start, '%d/%m/%Y').strftime("%B").capitalize()
        selected_year = datetime.strptime(start, '%d/%m/%Y').strftime("%Y")
        sub_heading = "Commissions dues par votre organisme au titre des crédits débloqués pendant le mois de" + \
            " "+selected_month+" , " + selected_year
        ws.cell(row=12, column=2).value = sub_heading
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for i, demande in enumerate(demandes):
            prospect = await Prospect.find_one(Prospect.prospect_id == demande.prospect_id)
            row = start_row + i
            ws.insert_rows(row)
            ws.cell(row=row, column=2).value = prospect.nom + \
                " " + prospect.prenom
            ws.cell(row=row, column=4).value = await CreditService.change_string_float(demande.montant)
            ws.cell(row=row, column=5).value = await CreditService.change_string_float(demande.hb)
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border

        wb.save("app/temp/facture_sofac.xlsx")

    @staticmethod
    async def generate_cdm_facture(start, end, bank):
        wb = openpyxl.load_workbook("app/static/facture_cdm.xlsx")
        ws = wb["Sheet1"]
        start_row = 14
        demandes = await Credit.find_many({"banque": "CDM", "statusCredit": "Débloqué", "status_honnaires.hb": False, "date_debloque": {
            "$gte": datetime.strptime(start, '%d/%m/%Y'),
            "$lte": datetime.strptime(end, '%d/%m/%Y'),
        }}).to_list()
        today_date = datetime.now().strftime("%m/%d/%Y")
        ws.cell(row=10, column=9).value = today_date
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        selected_month = datetime.strptime(start, '%d/%m/%Y').strftime("%B").capitalize()
        selected_year = datetime.strptime(start, '%d/%m/%Y').strftime("%Y")
        sub_heading = "Commissions dues par votre organisme au titre des crédits débloqués pendant le mois de" + \
            " "+selected_month+" , " + selected_year
        ws.cell(row=12, column=2).value = sub_heading
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.insert_rows(start_row, amount=demandes.__len__())
        for i, demande in enumerate(demandes):
            prospect = await Prospect.find_one(Prospect.prospect_id == demande.prospect_id)
            row = start_row + i

            ws.cell(row=row, column=2).value = prospect.nom + \
                " " + prospect.prenom
            ws.cell(row=row, column=3).value = prospect.cin_sejour
            ws.cell(row=row, column=5).value = prospect.profession
            ws.cell(row=row, column=6).value = demande.prospect_revenue
            ws.cell(row=row, column=8).value = await CreditService.change_string_float(demande.montant_debloque)
            ws.cell(row=row, column=7).value = demande.date_debloque.strftime("%B").capitalize()
            ws.cell(row=row, column=9).value = await CreditService.change_string_float(demande.hb)

            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7).border = thin_border
            ws.cell(row=row, column=8).border = thin_border
            ws.cell(row=row, column=9).border = thin_border

        wb.save("app/temp/facture_cdm.xlsx")

    @staticmethod
    async def generate_bmci_facture(start, end, bank):
        wb = openpyxl.load_workbook("app/static/facture_bmci.xlsx")
        ws = wb["Sheet1"]
        start_row = 14
        demandes = await Credit.find_many({"banque": "BMCI", "statusCredit": "Débloqué", "status_honnaires.hb": False, "date_debloque": {
            "$gte": datetime.strptime(start, '%d/%m/%Y'),
            "$lte": datetime.strptime(end, '%d/%m/%Y'),
        }}).to_list()
        if demandes.__len__() == 0 :
            raise HTTPException(status_code=404, detail="Aucun enregistrement trouvé")
        today_date = datetime.now().strftime("%m/%d/%Y")
        ws.cell(row=10, column=6).value = today_date
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        selected_month = datetime.strptime(start, '%d/%m/%Y').strftime("%B").capitalize()
        selected_year = datetime.strptime(start, '%d/%m/%Y').strftime("%Y")
        sub_heading = "Commissions dues par votre organisme au titre des crédits débloqués pendant le mois de" + \
            " "+selected_month+" , " + selected_year
        ws.cell(row=12, column=2).value = sub_heading
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.insert_rows(start_row, amount=demandes.__len__())
        for i, demande in enumerate(demandes):
            prospect = await Prospect.find_one(Prospect.prospect_id == demande.prospect_id)
            row = start_row + i
            
            ws.cell(row=row, column=2).value = prospect.nom + \
                " " + prospect.prenom
            ws.cell(row=row, column=3).value = prospect.cin_sejour
            bank = [b for b in demande.banque_envoye if b["banque"] == "BMCI"]
            agency = bank[0]['agence']
            if agency is None :
                 agency = ""
            ws.cell(row=row, column=4).value = bank[0]['agence']
            ws.cell(row=row, column=5).value = await CreditService.change_string_float(demande.montant)
            ws.cell(row=row, column=6).value = await CreditService.change_string_float(demande.hb)

            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6).border = thin_border
           
        wb.save("app/temp/facture_bmci.xlsx")

    @staticmethod
    async def generate_sgmb_facture(start, end, bank):
        wb = openpyxl.load_workbook("app/static/facture_sgmb.xlsx")
        ws = wb["Sheet1"]
        start_row = 14
        demandes = await Credit.find_many({"banque": "SGMB", "statusCredit": "Débloqué", "status_honnaires.hb": False, "date_debloque": {
            "$gte": datetime.strptime(start, '%d/%m/%Y'),
            "$lte": datetime.strptime(end, '%d/%m/%Y'),
        }}).to_list()
        if demandes.__len__() == 0 :
            raise HTTPException(status_code=404, detail="Aucun enregistrement trouvé")
        today_date = datetime.now().strftime("%m/%d/%Y")
        ws.cell(row=10, column=7).value = today_date
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        selected_month = datetime.strptime(start, '%d/%m/%Y').strftime("%B").capitalize()
        selected_year = datetime.strptime(start, '%d/%m/%Y').strftime("%Y")
        sub_heading = "Commissions dues par votre organisme au titre des crédits débloqués pendant le mois de" + \
            " "+selected_month+" , " + selected_year
        ws.cell(row=12, column=2).value = sub_heading
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.insert_rows(start_row, amount=demandes.__len__())
        for i, demande in enumerate(demandes):
            prospect = await Prospect.find_one(Prospect.prospect_id == demande.prospect_id)
            row = start_row + i
            
            ws.cell(row=row, column=2).value = prospect.nom + " " + prospect.prenom
            ws.cell(row=row, column=3).value = prospect.cin_sejour
            bank = [b for b in demande.banque_envoye if b["banque"] == "SGMB"]
            agency = bank[0]['agence']
            if agency is None :
                 agency = ""
            ws.cell(row=row, column=4).value = bank[0]['agence']
            ws.cell(row=row, column=5).value = await CreditService.change_string_float(demande.montant_valide)
            ws.cell(row=row, column=6).value = await CreditService.change_string_float(demande.montant_debloque)
            ws.cell(row=row, column=7).value = await CreditService.change_string_float(demande.hb)

            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7).border = thin_border
            
           
        wb.save("app/temp/facture_sgmb.xlsx")


    @staticmethod
    async def generate_bp_facture(start, end, bank):
        wb = openpyxl.load_workbook("app/static/facture_bp.xlsx")
        ws = wb["Sheet1"]
        start_row = 14
        demandes = await Credit.find_many({"banque": "BP", "statusCredit": "Débloqué", "status_honnaires.hb": False, "date_debloque": {
            "$gte": datetime.strptime(start, '%d/%m/%Y'),
            "$lte": datetime.strptime(end, '%d/%m/%Y'),
        }}).to_list()
        if demandes.__len__() == 0 :
            raise HTTPException(status_code=404, detail="Aucun enregistrement trouvé")
        today_date = datetime.now().strftime("%m/%d/%Y")
        ws.cell(row=10, column=7).value = today_date
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        selected_month = datetime.strptime(start, '%d/%m/%Y').strftime("%B").capitalize()
        selected_year = datetime.strptime(start, '%d/%m/%Y').strftime("%Y")
        sub_heading = "Commissions dues par votre organisme au titre des crédits débloqués pendant le mois de" + \
            " "+selected_month+" , " + selected_year
        ws.cell(row=12, column=2).value = sub_heading
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.insert_rows(start_row, amount=demandes.__len__())
        for i, demande in enumerate(demandes):
            prospect = await Prospect.find_one(Prospect.prospect_id == demande.prospect_id)
            row = start_row + i
            
            ws.cell(row=row, column=2).value = prospect.nom + " " + prospect.prenom
            ws.cell(row=row, column=3).value = prospect.cin_sejour
            bank = [b for b in demande.banque_envoye if b["banque"] == "BP"]
            agency = bank[0]['agence']
            if agency is None :
                 agency = ""
            ws.cell(row=row, column=4).value = bank[0]['agence']
            ws.cell(row=row, column=5).value = await CreditService.change_string_float(demande.montant_valide)
            ws.cell(row=row, column=6).value = await CreditService.change_string_float(demande.montant_debloque)
            ws.cell(row=row, column=7).value = await CreditService.change_string_float(demande.hb)

            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=6).border = thin_border
            ws.cell(row=row, column=7).border = thin_border
            
           
        wb.save("app/temp/facture_bp.xlsx")

    
    @staticmethod
    async def create_consommation_mandat(credit_id: UUID):
        doc = DocxTemplate("app/static/mandat_conso.docx")
        credit = await CreditService.get_credit_by_id(credit_id)
        prospect = await ProspectService.get_prospect_by_id(credit.prospect_id)
        if credit.montant_valide == "":
            montant = credit.montant
            montant_calcule = await CreditService.change_string_float(credit.montant)
        else:
            montant = credit.montant_valide
            montant_calcule = await CreditService.change_string_float(credit.montant_valide)
        frais = await CreditService.switch_mandat(montant_calcule)
        frais = await CreditService.change_float_string(frais)
        prospect_info = {
            "prospect_nom": prospect.nom + " " + prospect.prenom,
            "cin_sejour": prospect.cin_sejour,
        }

        credit_info = {
            "montant": montant,
            "type_credit": credit.type_credit,
            "frais": frais,
            'date': date.today().strftime('%d/%m/%Y')
        }
        context = {"prospect_info": prospect_info, "credit_info": credit_info}
        doc.render(context)
        fileLocation = "app/temp/"+str(credit_id)+'_mandat.docx'
        doc.save(fileLocation)

    @staticmethod
    async def create_autres_mandat(credit_id: UUID):
        doc = DocxTemplate("app/static/mandat.docx")
        credit = await CreditService.get_credit_by_id(credit_id)
        prospect = await ProspectService.get_prospect_by_id(credit.prospect_id)

        if credit.montant_valide == "":
            montant = credit.montant
            montant_calcule = await CreditService.change_string_float(credit.montant)
        else:
            montant = credit.montant_valide
            montant_calcule = await CreditService.change_string_float(credit.montant_valide)
        if (credit.type_credit) == "hypothecaire":
            doc = DocxTemplate("app/static/mandat_hypo.docx")
            frais = (5/100) * montant_calcule
        else:
            doc = DocxTemplate("app/static/mandat_immo.docx")
            frais = (1/100) * montant_calcule

        frais = await CreditService.change_float_string(frais)
        prospect_info = {
            "prospect_nom": prospect.nom + " " + prospect.prenom,
            "cin_sejour": prospect.cin_sejour,
            "adresse": prospect.adresse["adresse1"] + " " + prospect.adresse["ville"] + ", " + prospect.adresse["pays"]
        }

        credit_info = {
            "montant": montant,
            "type_credit": credit.type_credit.upper(),
            "frais": frais,
            'date': date.today().strftime('%d/%m/%Y')
        }
        context = {"prospect_info": prospect_info, "credit_info": credit_info}
        doc.render(context)
        fileLocation = "app/temp/"+str(credit_id)+'_mandat.docx'
        doc.save(fileLocation)

    @staticmethod
    async def download_mandat(request: Request, credit_id: UUID, type_credit: str):
        if (type_credit == "consommation"):
            await CreditService.create_consommation_mandat(credit_id)
        else:
            await CreditService.create_autres_mandat(credit_id)

        fileLocation = "app/temp/"+str(credit_id)+'_mandat.docx'
        return FileResponse(fileLocation, media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

    @staticmethod
    async def delete_credit(id: UUID):
        credit = await Credit.find_one(Credit.credit_id == id)
        if not credit:
            raise pymongo.errors.OperationFailure("Credit not Found")
        try:
            prospect = await Prospect.find_one(Prospect.prospect_id == credit.prospect_id)
            await prospect.update({
                "$pull": {"credits": credit.credit_id}
            })
            await credit.delete()

        except:
            return {
                "message": "Credit Not Deleted !"
            }
        return {
            "message": "Credit deleted successfully"
        }

    @staticmethod
    async def switch_hc(montant: float):
        frais = 0
        if montant > 300000:
            frais = (3/100)*montant
        elif montant <= 20000:
            frais = 1000
            return frais
        elif montant <= 30000:
            frais = 1500
            return frais
        elif montant <= 40000:
            frais = 2000
            return frais
        elif montant <= 50000:
            frais = 3000
            return frais
        elif montant <= 75000:
            frais = 3500
            return frais
        elif montant <= 100000:
            frais = 4000
            return frais
        elif montant <= 300000:
            frais = 1000
            return frais

    @staticmethod
    async def bulk_update():
        credits = await Credit.find_all().to_list()
        for credit in credits:
            prospect = await Prospect.find_one(Prospect.prospect_id == credit.prospect_id)
            credit.agent_id = prospect.agent_id
            await credit.save()
        return credits

    @staticmethod
    async def get_credits(user) -> List[CreditDisplay]:
        credits_obj = await Credit.find_all().aggregate([{
            "$lookup": {
                "from": "prospects",
                "let":  {"prospect_id": "$prospect_id"},
                "pipeline": [
                    {"$match": {
                        "$expr": {"$eq": ["$$prospect_id", "$prospect_id"]}}},
                ],
                "as": "prospectInfo"
            }
        },
            {
            "$unwind": "$prospectInfo"
        },
         {"$lookup": {

                        "from": "users",
                        "localField": "agent_id",
                        "foreignField": "user_id",
                        "as": "agentInfo",
                    }},
                    {
                        "$unwind": "$agentInfo"
                    },
            {
            "$project": {
                "prospectInfo": {
                    "nom": 1,
                    "prenom": 1,
                    "cin_sejour": 1,
                    "type_profession": 1,
                    "caisse": 1,
                    "renseignements_bancaires": 1,
                },
                "agentInfo": {
                        "user_name": 1,
                        "email": 1,
                    },
                "adresse_bien": 1,
                "banque": 1,
                "commentaires": 1,
                "credit_id": 1,
                "duree_credit": 1,
                "etat_bien": 1,
                "franchise": 1,
                "frequence": 1,
                "garanties": 1,
                "mensualite": 1,
                "montant": 1,
                "montant_acte": 1,
                "montant_debloque": 1,
                "montant_valide": 1,
                "montant_travaux": 1,
                "montant_venal": 1,
                "nature_bien": 1,
                "objet_credit": 1,
                "promoteur": 1,
                "promoteur_nom": 1,
                "prospect_id": 1,
                "qot_financement": 1,
                "statusCredit": 1,
                "superficie": 1,
                "taux": 1,
                "taux_demande": 1,
                "taux_endt": 1,
                "teg": 1,
                "titre_foncier": 1,
                "type_credit": 1,
                "usage_bien": 1,
                "date": 1,
                "hc": 1,
                "hb": 1,
                "status_honnaires": 1,
                "coemp": 1,
                "engagements_bancaires": 1,
                "prospect_revenue": 1,
                "banque_envoye": 1,
                "date_debloque": 1,
                "agent_id": 1,

            }
        }]).to_list()

        if user.role == 'Agent':
            newList = []
            for record in credits_obj:
                if record["agentInfo"]["email"] == user.email:
                    newList.append(record)
            credits_obj = newList

        return credits_obj

    @staticmethod
    async def get_credit_by_id(id: UUID) -> Optional[CreditOut]:
        credit = await Credit.find_one(Credit.credit_id == id)
        return credit

    @staticmethod
    async def get_credits_bulk_update(prospect_id, agent_id):
        credit = await Credit.find_one(Credit.credit_id == prospect_id)
        user = await UserService.get_user_by_id(agent_id)
        await credit.update({"$set": {"agent_id": user.user_id}})
        return credit
    
    @staticmethod
    async def update_agent_update(credit_id, agent_id):
        credit = await Credit.find_one(Credit.credit_id == credit_id)
        user = await UserService.get_user_by_id(agent_id)
        await credit.update({"$set": {"agent_id": user.user_id}})
        return credit

    @staticmethod
    async def get_prospect_record(id: UUID):

        credit = await Credit.find_one(Credit.credit_id == id)
        prospect = await Prospect.find_one(Prospect.prospect_id == credit.prospect_id)
        co_emp = None
        if prospect.coemp_id is not None:
            co_emp = await CoempService.get_coemp_by_id(prospect.coemp_id)
        record = {
            'prospect': prospect,
            'co_emp': co_emp,
            'credit': credit
        }
        if not Prospect:
            raise pymongo.errors.OperationFailure("Prospect not found")

        return record

    @staticmethod
    async def prospect_credits(id: UUID):
        prospect = await Prospect.find_one(Prospect.prospect_id == id)
        credits = await Credit.find(
            In(Credit.credit_id, prospect.credits)
        ).to_list()
        return credits

    @staticmethod
    async def update_record(payLoad: ProspectRecord):
        await Prospect.find_one(Prospect.prospect_id == payLoad.prospect.prospect_id).update({'$set': payLoad.prospect})
        if payLoad.credit is not None:
            await Credit.find_one(Credit.credit_id == payLoad.credit.credit_id).update({'$set': payLoad.credit})
        return payLoad

    @staticmethod
    async def demande_credit(data: DemandeCreditCreate):

        try:
            prospect_in = Prospect(
                prospect_id=data.prospect.prospect_id,
                nom=data.prospect.nom,
                prenom=data.prospect.prenom,
                datenaissance=data.prospect.datenaissance,
                lieunaissance=data.prospect.lieunaissance,
                nationalite=data.prospect.nationalite,
                adresse=data.prospect.adresse,
                telephone=data.prospect.telephone,
                situation=data.prospect.situation,
                profession=data.prospect.profession,
                telpro=data.prospect.telpro,
                datembauche=data.prospect.datembauche,
                revenue=data.prospect.revenue,
                coemp_id=data.prospect.coemp_id,
                agent_id=data.prospect.agent_id,
                renseignements_bancaires=data.prospect.renseignements_bancaires,
                engagements_bancaires=data.prospect.engagements_bancaires,
                credits=data.prospect.credits
            )

            credit_in = Credit(
                credit_id=data.credit.credit_id,
                type_credit=data.credit.type_credit,
                montant=data.credit.montant,
                duree_credit=data.credit.duree_credit,
                frequence=data.credit.frequence,
                mensualite=data.credit.mensualite,
                taux=data.credit.taux,
                franchise=data.credit.franchise,
                taux_endt=data.credit.taux_endt,
                teg=data.credit.teg,
                qot_financement=data.credit.qot_financement,
                objet_credit=data.credit.objet_credit,
                nature_bien=data.credit.nature_bien,
                etat_bien=data.credit.etat_bien,
                usage_bien=data.credit.usage_bien,
                montant_acte=data.credit.montant_acte,
                montant_venal=data.credit.montant_venal,
                adresse_bien=data.credit.adresse_bien,
                superficie=data.credit.superficie,
                statusCredit=data.credit.statusCredit,
                prospect_id=data.prospect.prospect_id
            )
            demande_credit_obj = DemandeCredit(
                prospect=prospect_in,
                credit=credit_in,
                statusDemande=data.statusDemande
            )

            await prospect_in.save()
            await credit_in.save()
            await demande_credit_obj.save()
        except Exception as msg:
            raise msg

        return demande_credit_obj
